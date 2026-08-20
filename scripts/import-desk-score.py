#!/usr/bin/env python3
"""Importa Score_Quantitativo do Excel enriquecido e cruza com editorial-ranking."""

from __future__ import annotations

import json
import re
import unicodedata
from datetime import datetime, timezone
from pathlib import Path

try:
    import openpyxl
except ImportError as exc:  # pragma: no cover
    raise SystemExit("Instale openpyxl: pip install openpyxl") from exc

ROOT = Path(__file__).resolve().parents[1]
XLSX = ROOT / "Imprensa" / "ranking_imprensa_nordeste_2026_enriquecido.xlsx"
EDITORIAL = ROOT / "data" / "editorial-ranking-v1.json"
OUT = ROOT / "data" / "desk-score-v1.json"


def norm(s: str | None) -> str:
    if not s:
        return ""
    s = str(s).upper()
    s = unicodedata.normalize("NFKD", s)
    s = "".join(c for c in s if not unicodedata.combining(c))
    s = re.sub(r"[^A-Z0-9]+", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    for prefix in ("PORTAL ", "BLOG ", "RADIO ", "SITE ", "JORNAL "):
        if s.startswith(prefix):
            s = s[len(prefix) :]
    return s


def num(v):
    if v is None or v == "":
        return None
    try:
        n = float(v)
        return n if n == n else None
    except (TypeError, ValueError):
        return None


def main() -> None:
    if not XLSX.exists():
        raise SystemExit(f"Arquivo não encontrado: {XLSX}")
    editorial = json.loads(EDITORIAL.read_text(encoding="utf-8"))
    by_uf_rank = {(i["uf"], int(i["rank"])): i for i in editorial["items"]}
    by_uf_name = {(i["uf"], norm(i["name"])): i for i in editorial["items"]}

    wb = openpyxl.load_workbook(XLSX, data_only=True)
    ws = wb["Score_Quantitativo"]
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]

    dashboard = []
    for row in wb["Dashboard_Score"].iter_rows(min_row=3, max_row=11, values_only=True):
        if row[0] and len(str(row[0])) == 2:
            dashboard.append(
                {
                    "uf": row[0],
                    "state": row[1],
                    "top20": row[2],
                    "withFollowers": row[3],
                    "withReach": row[4],
                    "coverageAtLeastMedium": row[5],
                    "avgScore": row[6],
                    "quantitativeLeader": row[7],
                }
            )

    rules = []
    for row in wb["Dashboard_Score"].iter_rows(min_row=13, max_row=20, values_only=True):
        if row[0] and row[1]:
            rules.append({"component": str(row[0]), "rule": str(row[1])})

    items = []
    unmatched = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        d = dict(zip(headers, row))
        uf = d["UF"]
        rank = int(d["Rank editorial"])
        name = d["Veículo"]
        ed = by_uf_rank.get((uf, rank)) or by_uf_name.get((uf, norm(name)))
        if not ed:
            unmatched.append({"uf": uf, "rank": rank, "name": name})
            continue

        followers = num(d.get("Seguidores Instagram"))
        reach = num(d.get("Métrica alcance/audiência"))
        evidence = num(d.get("Qualidade evidência (0-10)"))
        score_final = num(d.get("Score final (0-100)"))
        items.append(
            {
                "vehicleId": ed["vehicleId"],
                "uf": uf,
                "state": d.get("Estado"),
                "name": name,
                "type": ed.get("type") or d.get("Tipo"),
                "city": d.get("Cidade"),
                "editorialRank": rank,
                "quantitativeRank": int(d["Rank quantitativo"]) if d.get("Rank quantitativo") is not None else None,
                "deskFollowers": int(followers) if followers is not None else None,
                "deskReachValue": reach,
                "deskReachUnit": d.get("Unidade alcance") or None,
                "deskMetricSource": d.get("Fonte métrica") or None,
                "deskSourceType": d.get("Tipo da fonte") or None,
                "deskEvidenceQuality": evidence,
                "deskObservation": d.get("Observação") or None,
                "deskScoreEditorial": num(d.get("Score editorial (0-50)")),
                "deskScoreDigital": num(d.get("Score digital (0-25)")),
                "deskScoreReach": num(d.get("Score alcance (0-15)")),
                "deskScoreEvidence": num(d.get("Score evidência (0-10)")),
                "deskScoreFinal": score_final,
                "deskCoverage": d.get("Cobertura de dados") or None,
            }
        )

    with_followers = sum(1 for i in items if i["deskFollowers"] is not None)
    with_reach = sum(1 for i in items if i["deskReachValue"] is not None)
    coverage = {}
    for i in items:
        coverage[i["deskCoverage"] or "?"] = coverage.get(i["deskCoverage"] or "?", 0) + 1

    out = {
        "version": "desk-score-v1",
        "sourceFile": "Imprensa/ranking_imprensa_nordeste_2026_enriquecido.xlsx",
        "sheet": "Score_Quantitativo",
        "importedAt": datetime.now(timezone.utc).isoformat(),
        "rules": rules,
        "dashboard": dashboard,
        "validation": {
            "total": len(items),
            "unmatched": len(unmatched),
            "withFollowers": with_followers,
            "withReach": with_reach,
            "coverage": coverage,
            "ok": len(unmatched) == 0 and len(items) == 180,
        },
        "unmatched": unmatched,
        "items": items,
    }
    OUT.write_text(json.dumps(out, ensure_ascii=False, indent=2), encoding="utf-8")
    v = out["validation"]
    print(
        f"OK {OUT} matched={v['total']} unmatched={v['unmatched']} "
        f"followers={v['withFollowers']} reach={v['withReach']} ok={v['ok']}"
    )


if __name__ == "__main__":
    main()
