#!/usr/bin/env python3
"""Importa Ranking_20 do levantamento humano e cruza com vehicles-v1.json."""

from __future__ import annotations

import json
import re
import unicodedata
from datetime import datetime, timezone
from pathlib import Path
from urllib.parse import urlparse

try:
    import openpyxl
except ImportError as exc:  # pragma: no cover
    raise SystemExit("Instale openpyxl: pip install openpyxl") from exc

ROOT = Path(__file__).resolve().parents[1]
XLSX = ROOT / "Imprensa" / "ranking_imprensa_nordeste_2026.xlsx"
VEHICLES = ROOT / "data" / "vehicles-v1.json"
OUT = ROOT / "data" / "editorial-ranking-v1.json"


def norm(s: str | None) -> str:
    if not s:
        return ""
    s = str(s).upper()
    s = unicodedata.normalize("NFKD", s)
    s = "".join(c for c in s if not unicodedata.combining(c))
    s = re.sub(r"[^A-Z0-9]+", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    for prefix in ("PORTAL ", "BLOG ", "RADIO ", "SITE ", "JORNAL "):
        if s.startswith(prefix):
            s = s[len(prefix) :]
    return s


def host(url: str | None) -> str:
    if not url:
        return ""
    try:
        raw = str(url)
        u = raw if raw.startswith("http") else f"https://{raw}"
        h = urlparse(u).hostname or ""
        return h.lower().replace("www.", "")
    except Exception:
        return ""


def want_type(raw) -> str:
    r = str(raw or "")
    if "Portal" in r:
        return "Portal"
    if "Jornal" in r:
        return "Jornal"
    if "Blog" in r:
        return "Blog"
    if re.search(r"R.?dio", r, re.I):
        return "Rádio"
    return r


def band_from(just: str | None) -> str | None:
    m = re.search(r"Faixa\s+([A-C](?:/[A-C])?)", str(just or ""), re.I)
    return m.group(1).upper() if m else None


def pick(cands: list[dict], entry: dict) -> dict | None:
    if not cands:
        return None
    if len(cands) == 1:
        return cands[0]
    exact = [
        c
        for c in cands
        if norm(c["name"]) == norm(entry["name"]) or c["name"].upper() == entry["name"].upper()
    ]
    if len(exact) == 1:
        return exact[0]
    h = host(entry.get("sourceDirectory"))
    if h:
        by_h = [c for c in cands if host(c.get("website")) == h]
        if by_h:
            return by_h[0]
    typed = [c for c in cands if c["type"] == entry["type"]]
    return (typed or cands)[0]


def main() -> None:
    if not XLSX.exists():
        raise SystemExit(f"Arquivo não encontrado: {XLSX}")

    wb = openpyxl.load_workbook(XLSX, data_only=True)
    ws = wb["Ranking_20"]
    editorial_rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        editorial_rows.append(
            {
                "uf": row[0],
                "state": row[1],
                "rank": int(row[2]),
                "name": row[3],
                "type": want_type(row[4]),
                "city": row[5],
                "justification": row[6],
                "sourceDirectory": row[7],
                "sourceDesk": row[8],
                "confidence": row[9],
                "band": band_from(row[6]),
            }
        )

    methodology = []
    for row in wb["Metodologia"].iter_rows(min_row=3, values_only=True):
        if row[0]:
            methodology.append({"key": row[0], "value": row[1]})

    vehicles = json.loads(VEHICLES.read_text(encoding="utf-8"))
    by_uf_name: dict[tuple[str, str], list] = {}
    by_host: dict[tuple[str, str], list] = {}
    for v in vehicles:
        by_uf_name.setdefault((v["uf"], norm(v["name"])), []).append(v)
        h = host(v.get("website"))
        if h:
            by_host.setdefault((v["uf"], h), []).append(v)

    matched = []
    unmatched = []
    for e in editorial_rows:
        key = (e["uf"], norm(e["name"]))
        cands = by_uf_name.get(key, [])
        typed = [c for c in cands if c["type"] == e["type"]]
        chosen = pick(typed or cands, e)
        if not chosen:
            chosen = pick(by_host.get((e["uf"], host(e.get("sourceDirectory"))), []), e)
        if not chosen:
            n = norm(e["name"])
            fuzzy = [
                v
                for v in vehicles
                if v["uf"] == e["uf"]
                and v["type"] == e["type"]
                and (n in norm(v["name"]) or norm(v["name"]) in n)
            ]
            chosen = pick(fuzzy, e)
        if not chosen:
            unmatched.append(e)
            continue
        matched.append(
            {
                **e,
                "vehicleId": chosen["id"],
                "matchedName": chosen["name"],
                "matchedType": chosen["type"],
                "matchMethod": "name+type" if typed else "fallback",
            }
        )

    out = {
        "version": "editorial-v1",
        "sourceFile": "Imprensa/ranking_imprensa_nordeste_2026.xlsx",
        "importedAt": datetime.now(timezone.utc).isoformat(),
        "methodology": methodology,
        "rules": {
            "scope": "Top 20 veículos não-TV por estado (mistos)",
            "tvPolicy": "TV fora do quantitativo de 20; listar separadamente",
            "criteria": [
                "Audiência / circulação / seguidores com evidência pública",
                "Alcance estadual e capilaridade regional",
                "Tradição, força de marca e relevância editorial",
                "Presença digital e capacidade de pautar o debate",
            ],
        },
        "total": len(matched),
        "unmatched": unmatched,
        "items": matched,
    }
    OUT.write_text(json.dumps(out, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"OK {OUT} — matched={len(matched)} unmatched={len(unmatched)}")


if __name__ == "__main__":
    main()
